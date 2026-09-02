from datetime import date
from pathlib import Path
import ast
import json
import re
import sys

from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parent / 'vault-ia-local' if (Path(__file__).resolve().parent / 'vault-ia-local').exists() else Path(__file__).resolve().parent.parent
errors=[]; warnings=[]; justified=[]
def rel(p): return p.relative_to(ROOT).as_posix()
EXCLUDED_REPORTS={'VALIDACAO.md','VALIDACAO-COMPLETA.md'}
IGNORED_DIRS={'.git','.obsidian','.venv','venv','__pycache__'}
TEXT_EXTENSIONS={'.md','.py','.json','.txt','.yml','.yaml','.ps1','.toml','.ini','.cfg','.sh','.pem','.key','.properties'}
SPECIAL_TEXT_PREFIXES=('.env',)

def ignored(p):
    return any(part in IGNORED_DIRS for part in p.relative_to(ROOT).parts)

all_md=[p for p in ROOT.rglob('*.md') if not ignored(p)]
md=[p for p in all_md if p.name not in EXCLUDED_REPORTS]
text_files=[p for p in ROOT.rglob('*') if p.is_file() and not ignored(p)
            and (p.suffix.lower() in TEXT_EXTENSIONS or p.name.lower().startswith(SPECIAL_TEXT_PREFIXES))]
JUST=re.compile(r'<!--\s*validador:\s*(sem-referencias|sem-data|revisao-vencida)\s*:\s*(.+?)\s*-->')
NO_FORMULAS_JUST=re.compile(r'<!--\s*validador:\s*sem-formulas\s*:\s*(.+?)\s*-->')

# Nota de controle: frontmatter YAML com control_id ativa o schema de controle
# (cadeia risco → requisito → configuração → teste → evidência → operação →
# revisão). Campos e seções exigidos estão em
# 99-Templates/Modelo-de-nota-de-controle.md.
FRONTMATTER=re.compile(r'\A---\s*\n(.*?)\n---\s*\n',re.S)
ISO_DATE=re.compile(r'\d{4}-\d{2}-\d{2}')
CONTROL_STATUSES={'documented','configured','tested','verified'}
CONTROL_EVIDENCE={'manufacturer-specification','own-test','inference','estimate','editorial','compensating-control'}
CONTROL_SECTIONS=('## Risco','## Configuração','## Teste positivo','## Teste negativo','## Evidência','## Rollback')
control_notes=0

for p in md:
    text=p.read_text(encoding='utf-8', errors='replace')
    # Reference URL syntax, without network access.
    for url in re.findall(r'\]\:\s*(https?://\S+)', text):
        if any(ch in url for ch in ['<','>','"']):
            warnings.append(f'MALFORMED_REFERENCE_URL {rel(p)}')
    # Internal Obsidian links.
    for raw in re.findall(r'\[\[([^\]|#]+)', text):
        target=raw.strip()
        candidates=[ROOT/(target+'.md'), ROOT/target]
        if not any(c.exists() for c in candidates):
            errors.append(f'LINK {rel(p)} -> {target}')
    # Heuristic markers for accidental agent traces. This is a consistency
    # check, not a complete prompt-injection detector.
    if re.search(r'Need update|Use file edit|tool_result_received|compacted_history|functions\.(file|shell|plan|message)|<system_reminder>', text, re.I):
        errors.append(f'AGENT_TRACE {rel(p)}')
    if r'\\n\\n' in text or r'\\n' in text:
        errors.append(f'ESCAPED_NEWLINE {rel(p)}')
    # Absolute paths in published Markdown.
    if re.search(r'[A-Za-z]:\\|/home/ubuntu/|/Users/|C:\\Users', text):
        errors.append(f'ABSOLUTE_PATH {rel(p)}')
    # Editorial completeness checks. A note may justify an absence with
    # <!-- validador: sem-referencias: motivo --> or <!-- validador: sem-data: motivo -->.
    just={k:v for k,v in JUST.findall(text)}
    if len(text.splitlines()) >= 25 and '## Referências' not in text and '## References' not in text:
        if 'sem-referencias' in just: justified.append(f'NO_REFERENCES_SECTION {rel(p)} — {just["sem-referencias"]}')
        else: warnings.append(f'NO_REFERENCES_SECTION {rel(p)}')
    if len(text.splitlines()) >= 25 and not re.search(r'20\d{2}|Última atualização|Data|Status', text, re.I):
        if 'sem-data' in just: justified.append(f'NO_DATE_OR_STATUS {rel(p)} — {just["sem-data"]}')
        else: warnings.append(f'NO_DATE_OR_STATUS {rel(p)}')
    # Schema de nota de controle. status tested/verified exige teste próprio;
    # revisão vencida vira aviso (justificável com revisao-vencida) para o
    # gate estrito cobrar a re-verificação sem quebrar builds históricos.
    fm=FRONTMATTER.match(text)
    if fm and re.search(r'^control_id\s*:',fm.group(1),re.M):
        control_notes+=1
        front=fm.group(1)
        def field(name,front=front):
            # [ \t] e não \s: \s casaria o \n de um valor vazio e capturaria a linha seguinte.
            m=re.search(rf'^{name}[ \t]*:[ \t]*(.+?)[ \t]*$',front,re.M)
            return m.group(1).strip().strip('"\'') if m else None
        cid=field('control_id')
        if not cid or not re.fullmatch(r'[A-Z0-9]+(?:-[A-Z0-9]+){1,4}',cid):
            errors.append(f'CONTROL_BAD_ID {rel(p)}:{cid}')
        status=field('status')
        if status not in CONTROL_STATUSES:
            errors.append(f'CONTROL_BAD_STATUS {rel(p)}:{status}')
        evidence=field('evidence_type')
        if evidence not in CONTROL_EVIDENCE:
            errors.append(f'CONTROL_BAD_EVIDENCE_TYPE {rel(p)}:{evidence}')
        if status in {'tested','verified'} and evidence!='own-test':
            errors.append(f'CONTROL_STATUS_EVIDENCE_MISMATCH {rel(p)}: status={status} exige evidence_type=own-test')
        for date_field in ('verified_on','review_due'):
            value=field(date_field)
            if not value or not ISO_DATE.fullmatch(value):
                errors.append(f'CONTROL_BAD_DATE {rel(p)}:{date_field}={value}')
        due=field('review_due')
        if due and ISO_DATE.fullmatch(due) and due < date.today().isoformat():
            if 'revisao-vencida' in just:
                justified.append(f'CONTROL_REVIEW_OVERDUE {rel(p)}:{due} — {just["revisao-vencida"]}')
            else:
                warnings.append(f'CONTROL_REVIEW_OVERDUE {rel(p)}:{due}')
        if not field('risk'):
            errors.append(f'CONTROL_MISSING_RISK {rel(p)}')
        for section in CONTROL_SECTIONS:
            if section not in text:
                errors.append(f'CONTROL_MISSING_SECTION {rel(p)}:{section}')

# Python syntax.
for p in ROOT.rglob('*.py'):
    if ignored(p): continue
    try: ast.parse(p.read_text(encoding='utf-8'), filename=str(p))
    except SyntaxError as e: errors.append(f'PYTHON {rel(p)}:{e.lineno}:{e.msg}')

# 3) Dependency pinning and minimum safe versions used by the RAG example.
req=ROOT/'07-Implementacao-Casa/requirements-rag.txt'
req_pins={}
if not req.exists():
    errors.append('MISSING_RAG_REQUIREMENTS')
else:
    for i,line in enumerate(req.read_text(encoding='utf-8').splitlines(),1):
        line=line.strip()
        if not line or line.startswith('#'):
            continue
        match=re.fullmatch(r'([A-Za-z0-9_.-]+)==([^;\s]+)(?:\s*;.*)?', line)
        if not match:
            errors.append(f'UNPINNED_DEPENDENCY requirements-rag.txt:{i}:{line}')
            continue
        package=match.group(1).lower().replace('_','-')
        if package in req_pins:
            errors.append(f'DUPLICATE_DIRECT_DEPENDENCY requirements-rag.txt:{i}:{package}')
        req_pins[package]=match.group(2)

def parsed_version(value):
    match=re.match(r'^(\d+)\.(\d+)\.(\d+)', value or '')
    return tuple(map(int,match.groups())) if match else None

for package in {'chromadb','requests'}:
    if package in req_pins:
        # The example uses exact in-memory retrieval and urllib from the standard
        # library, so neither package belongs in its direct requirements. Some
        # model libraries may still bring requests transitively; SCA checks the lock.
        errors.append(f'RAG_FORBIDDEN_DIRECT_DEPENDENCY {package}')
for package in {'pypdf','sentence-transformers','huggingface-hub','numpy'}:
    if package not in req_pins:
        errors.append(f'RAG_REQUIRED_DIRECT_DEPENDENCY {package}')
for package,minimum in {'pypdf':(6,15,0)}.items():
    current=parsed_version(req_pins.get(package))
    if current is None:
        errors.append(f'RAG_REQUIRED_DEPENDENCY {package}')
    elif current < minimum:
        errors.append(f'RAG_DEPENDENCY_BELOW_SAFE_FLOOR {package}=={req_pins[package]} < {".".join(map(str,minimum))}')

# O checksum prova que o lock não mudou silenciosamente; esta reconciliação
# prova que ele ainda contém exatamente os pins diretos declarados acima.
lock=ROOT/'07-Implementacao-Casa/requirements-rag.lock.txt'
if not lock.exists():
    errors.append('MISSING_RAG_LOCKFILE')
else:
    lock_pins={}
    for match in re.finditer(r'^([A-Za-z0-9_.-]+)==([^\s\\]+)',lock.read_text(encoding='utf-8'),re.M):
        package=match.group(1).lower().replace('_','-')
        if package in lock_pins:
            errors.append(f'DUPLICATE_LOCKED_DEPENDENCY {package}')
        lock_pins[package]=match.group(2)
    for package,version in sorted(req_pins.items()):
        locked=lock_pins.get(package)
        if locked is None:
            errors.append(f'RAG_DIRECT_PIN_MISSING_FROM_LOCK {package}=={version}')
        elif locked != version:
            errors.append(f'RAG_DIRECT_PIN_LOCK_MISMATCH {package}: requirements={version} lock={locked}')

# 4) Static consistency invariants for the hardened local RAG example. Runtime
# behavior is exercised separately by --selftest in CI; these string/AST checks
# are intentionally not presented as a security proof.
rag=ROOT/'07-Implementacao-Casa/RAG-local-executavel.py'
if not rag.exists():
    errors.append('MISSING_RAG_SCRIPT')
else:
    rag_text=rag.read_text(encoding='utf-8',errors='replace')
    try:
        rag_tree=ast.parse(rag_text,filename=str(rag))
        imported=set()
        for node in ast.walk(rag_tree):
            if isinstance(node,ast.Import):
                imported.update(alias.name.split('.')[0] for alias in node.names)
            elif isinstance(node,ast.ImportFrom) and node.module:
                imported.add(node.module.split('.')[0])
        for forbidden_import in {'chromadb','requests'}:
            if forbidden_import in imported:
                errors.append(f'RAG_FORBIDDEN_IMPORT {forbidden_import}')
        if re.search(r'\bPersistentClient\b',rag_text):
            errors.append('RAG_PERSISTENT_VECTOR_DB_FORBIDDEN')
    except SyntaxError:
        # The generic Python syntax gate below reports line-level details.
        pass
    required_flags={
        '--selftest','--retrieve-only','--top-k','--max-files','--max-entries',
        '--max-file-bytes','--max-chunks','--max-text-chars',
        '--max-pdf-pages','--pdf-timeout','--ollama-timeout',
        '--allow-remote-ollama',
    }
    for flag in sorted(required_flags):
        if flag not in rag_text:
            errors.append(f'RAG_REQUIRED_CLI_GUARD {flag}')
    if '.is_symlink(' not in rag_text:
        errors.append('RAG_REQUIRED_GUARD symlink-rejection')
    if '.relative_to(' not in rag_text:
        errors.append('RAG_REQUIRED_GUARD relative-source-metadata')

# 5) Static checks for the Ollama enterprise security controls. These checks
# validate that the control package remains present and internally consistent;
# they do not claim that a runtime test was executed by this validator.
ollama_note=ROOT/'08-Implementacao-Empresa/04-LM-Studio-e-Ollama-como-runtimes-internos.md'
ollama_evidence=ROOT/'10-Operacao-e-Seguranca/Evidencias/Ollama-testes-negativos-2026-09-01.md'
ollama_controls={
    'OLL-NET-001': ROOT/'10-Operacao-e-Seguranca/Controles/OLL-NET-001-Bind-loopback.md',
    'OLL-AUTH-002': ROOT/'10-Operacao-e-Seguranca/Controles/OLL-AUTH-002-Camada-externa-de-autenticacao.md',
    'OLL-CLD-003': ROOT/'10-Operacao-e-Seguranca/Controles/OLL-CLD-003-Modo-somente-local.md',
}
if not ollama_note.exists():
    errors.append('MISSING_OLLAMA_ENTERPRISE_RUNTIME_NOTE')
else:
    note_text=ollama_note.read_text(encoding='utf-8',errors='replace')
    for required in ('127.0.0.1:11434','OLLAMA_NO_CLOUD','OLLAMA_ORIGINS','Require Authentication','store:false'):
        if required not in note_text:
            errors.append(f'OLLAMA_NOTE_MISSING {required}')
for control_id,control_path in ollama_controls.items():
    if not control_path.exists():
        errors.append(f'MISSING_OLLAMA_CONTROL {control_id}')
if not ollama_evidence.exists():
    errors.append('MISSING_OLLAMA_SECURITY_EVIDENCE')
else:
    evidence_text=ollama_evidence.read_text(encoding='utf-8',errors='replace')
    for required in ('OLL-NET-001','OLL-AUTH-002','OLL-CLD-003','127.0.0.1:11434','OLLAMA_NO_CLOUD:false'):
        if required not in evidence_text:
            errors.append(f'OLLAMA_EVIDENCE_MISSING {required}')

# 6) XLSX formulas and broken references.
for p in ROOT.rglob('*.xlsx'):
    if ignored(p): continue
    try:
        wb=load_workbook(p, data_only=False)
        formulas=[]
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c.value,str) and c.value.startswith('='):
                        formulas.append((ws.title,c.coordinate,c.value))
                        if '#REF!' in c.value or '#DIV/0!' in c.value:
                            errors.append(f'FORMULA {p.name}!{ws.title}!{c.coordinate} {c.value}')
        if not formulas:
            sidecar=p.with_suffix('.md')
            match=(NO_FORMULAS_JUST.search(sidecar.read_text(encoding='utf-8',errors='replace'))
                   if sidecar.exists() else None)
            if match:
                justified.append(f'NO_FORMULAS {rel(p)} — {match.group(1)}')
            else:
                warnings.append(f'NO_FORMULAS {rel(p)}')
    except Exception as e: errors.append(f'XLSX {rel(p)} {e}')

# 7) Heuristic secret indicators across relevant text/config/script formats.
# Regexes catch common accidental disclosures but do not replace a dedicated
# secret scanner, entropy analysis, repository history review or human audit.
secret_indicator=re.compile(
    r'BEGIN (PRIVATE KEY|RSA PRIVATE KEY)'
    r'|sk-[A-Za-z0-9]{20,}'
    r'|OPENAI_API_KEY[ \t]*=[ \t]*[^$<\s]+'
    r'|api[_-]?key[ \t]*[:=][ \t]*[A-Za-z0-9_-]{20,}'
    r'|password[ \t]*[:=][ \t]*[^<\s]{8,}'
    r'|bearer\s+[A-Za-z0-9._-]{20,}',
    re.I,
)
for p in text_files:
    t=p.read_text(encoding='utf-8',errors='replace')
    if secret_indicator.search(t):
        errors.append(f'POSSIBLE_SECRET_PATTERN {rel(p)}')

# 8) Third-party GitHub Actions must be immutable. The adjacent version comment
# remains human-readable while the 40-character commit prevents tag drift.
action_use=re.compile(r'^\s*(?:-\s*)?uses:\s+([^\s#]+)',re.M)
for p in [*ROOT.rglob('*.yml'),*ROOT.rglob('*.yaml')]:
    if ignored(p): continue
    for action in action_use.findall(p.read_text(encoding='utf-8',errors='replace')):
        if action.startswith('./'):
            continue
        if not re.fullmatch(r'[^@\s]+@[0-9a-f]{40}',action):
            errors.append(f'UNPINNED_GITHUB_ACTION {rel(p)}:{action}')

report=ROOT/'VALIDACAO-COMPLETA.md'
lines=['# Validação automatizada do vault','',
       f'- Markdown no pacote: {len(all_md)}',
       f'- Markdown analisados: {len(md)} (excluídos os relatórios {", ".join(sorted(EXCLUDED_REPORTS))})',
       f'- Arquivos na triagem textual heurística: {len(text_files)} ({", ".join(sorted(TEXT_EXTENSIONS))}; nomes especiais: .env*)',
       f'- Notas de controle (frontmatter control_id): {control_notes}',
       '- Limite da triagem: regexes indicam padrões suspeitos; não substituem secret scanning dedicado, histórico Git ou revisão humana.',
       f'- Erros: {len(errors)}',f'- Avisos: {len(warnings)}',f'- Avisos justificados: {len(justified)}','', '## Erros']
lines += [f'- {x}' for x in errors] or ['- Nenhum erro.']
lines += ['', '## Avisos'] + ([f'- {x}' for x in warnings] or ['- Nenhum aviso.'])
lines += ['', '## Avisos justificados'] + ([f'- {x}' for x in justified] or ['- Nenhum.'])
report.write_text(chr(10).join(lines)+chr(10),encoding='utf-8')
print(json.dumps({'markdown_pacote':len(all_md),'markdown_analisados':len(md),'notas_controle':control_notes,'errors':len(errors),'warnings':len(warnings),'justified':len(justified)},ensure_ascii=False))
if errors: sys.exit(1)
if '--strict' in sys.argv and warnings:
    print('STRICT: avisos não justificados')
    sys.exit(1)
