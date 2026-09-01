#!/usr/bin/env python3
"""Confere a planilha TCO-local-vs-OpenAI.xlsx contra um modelo independente em Python.

- Verifica que toda referência de célula em fórmula aponta para célula preenchida.
- Recalcula TCO local, custo API e break-evens a partir de Premissas (modelo próprio).
- Se o arquivo tiver valores em cache (salvo pelo Excel/LibreOffice), compara-os
  com o modelo e exige Checks!STATUS GERAL = PASS.
Saída 0 = OK; 1 = divergência; 2 = arquivo sem valores em cache (recalcule com
99-Templates/recalcular_tco.ps1 ou abra e salve no Excel/LibreOffice).
"""
import re, sys
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / '09-Servicos-e-Custos/TCO-local-vs-OpenAI.xlsx'
REF = re.compile(r"(?:(\w+)!)?\$?([A-Z]{1,2})\$?(\d+)(?::\$?([A-Z]{1,2})\$?(\d+))?")

def refs_de(formula):
    """Gera (sheet, col, row) para referências simples; para intervalos, só os dois extremos."""
    for sh, c1, r1, c2, r2 in REF.findall(formula):
        yield sh, c1, r1
        if c2:
            yield sh, c2, r2

def premissas(ws):
    return {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(2, ws.max_row + 1) if ws.cell(r, 1).value}

def modelo(p, precos):
    fx = p['Câmbio efetivo']; capex = p['CAPEX local']; vida = p['Vida útil']; res = p['Valor residual']
    if p.get('Método de CAPEX') == 1:
        rm = (1 + p['Taxa de desconto anual']) ** (1 / 12) - 1
        capex_m = (capex - res / (1 + rm) ** vida) * rm / (1 - (1 + rm) ** (-vida))
    else:
        capex_m = (capex - res) / vida
    tco = capex_m + p['Potência média'] * p['Horas/dia'] * p['Dias/mês'] * p['Tarifa efetiva'] + p['Refrigeração mensal'] + p['Manutenção/garantia mensal'] + p['Operação/espaço mensal']
    tin, tca, tout = p['Tokens entrada'], p['Tokens cached'], p['Tokens saída']
    f_in = p.get('Sobretaxa contexto longo — entrada', 1); f_ca = p.get('Sobretaxa contexto longo — cached', 1); f_out = p.get('Sobretaxa contexto longo — saída', 1)
    cw = p.get('Cache writes', 0); tot = tin + tca + tout
    out = {'tco': tco}
    for nome, (pi, pc, po) in precos.items():
        brl = ((tin * pi * f_in) + (tca * (pc + cw) * f_ca) + (tout * po * f_out)) * fx
        out[nome] = {'api_brl': brl, 'break_even': tco / (brl / tot)}
    return out

def main():
    wb = load_workbook(XLSX)
    bad = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith('='):
                    for sh, col, r in refs_de(c.value):
                        tgt = wb[sh] if sh else ws
                        if tgt[f'{col}{r}'].value is None:
                            bad.append(f'{ws.title}!{c.coordinate} -> {sh or ws.title}!{col}{r}')
    if bad:
        print('Referências para células vazias:', *bad, sep='\n  '); return 1
    p = premissas(wb['Premissas'])
    api = wb['API_OpenAI']
    precos = {api.cell(r, 1).value: (api.cell(r, 2).value, api.cell(r, 3).value, api.cell(r, 4).value) for r in range(2, api.max_row + 1) if api.cell(r, 1).value}
    m = modelo(p, precos)
    print(f"Modelo Python: TCO local = {m['tco']:.2f} BRL/mês")
    for nome in precos:
        print(f"  {nome}: API {m[nome]['api_brl']:.2f} BRL/mês | break-even {m[nome]['break_even']:.2f} M tokens/mês")
    wv = load_workbook(XLSX, data_only=True)
    if wv['Local']['C7'].value is None:
        print('AVISO: planilha sem valores em cache; comparação com Excel não realizada.'); return 2
    diffs = []
    def cmp(label, got, exp, tol=0.01):
        if got is None or abs(got - exp) > tol: diffs.append(f'{label}: planilha={got} modelo={exp:.4f}')
    cmp('Local!C7', wv['Local']['C7'].value, m['tco'])
    for i, nome in enumerate(precos, start=2):
        cmp(f'API_OpenAI!F{i}', wv['API_OpenAI'][f'F{i}'].value, m[nome]['api_brl'])
        cmp(f'Break_even!E{i}', wv['Break_even'][f'E{i}'].value, m[nome]['break_even'])
    ck = wv['Checks']
    status = next((ck.cell(r, 5).value for r in range(2, ck.max_row + 1) if ck.cell(r, 1).value == 'STATUS GERAL'), None)
    if status != 'PASS': diffs.append(f'Checks!STATUS GERAL = {status}')
    if diffs:
        print('DIVERGÊNCIAS:', *diffs, sep='\n  '); return 1
    print('OK: valores em cache batem com o modelo Python; Checks = PASS'); return 0

if __name__ == '__main__':
    sys.exit(main())
